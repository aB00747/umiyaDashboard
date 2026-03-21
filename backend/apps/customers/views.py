from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse
from django.db.models import Q
import csv
import io
from openpyxl import Workbook, load_workbook
from .models import Customer, CustomerType
from .serializers import CustomerSerializer, CustomerListSerializer, CustomerTypeSerializer


class CustomerTypeViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = CustomerType.objects.all()
    serializer_class = CustomerTypeSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None


class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.select_related('customer_type').all()
    search_fields = ['first_name', 'last_name', 'company_name', 'email', 'phone', 'gstin']
    ordering_fields = ['first_name', 'last_name', 'company_name', 'created_at', 'city']
    filterset_fields = ['is_active', 'customer_type', 'state', 'city']

    def get_serializer_class(self):
        if self.action == 'list':
            return CustomerListSerializer
        return CustomerSerializer

    @action(detail=False, methods=['get'])
    def search(self, request):
        q = request.query_params.get('q', '')
        if len(q) < 2:
            return Response([])
        customers = Customer.objects.select_related('customer_type').filter(
            Q(first_name__icontains=q) | Q(last_name__icontains=q) |
            Q(company_name__icontains=q) | Q(email__icontains=q) | Q(phone__icontains=q)
        )[:20]
        return Response(CustomerListSerializer(customers, many=True).data)

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser], url_path='import')
    def import_customers(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        created = 0
        errors = []

        # Pre-build a map of customer type names to objects
        type_map = {ct.name: ct for ct in CustomerType.objects.all()}

        def resolve_type(type_name):
            if not type_name:
                return None
            type_name = str(type_name).strip()
            if type_name in type_map:
                return type_map[type_name]
            # Create on the fly if not found
            ct, _ = CustomerType.objects.get_or_create(name=type_name)
            type_map[type_name] = ct
            return ct

        try:
            if file.name.endswith(('.xlsx', '.xls')):
                wb = load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    row_data = dict(zip(headers, row))
                    try:
                        Customer.objects.create(
                            first_name=row_data.get('first_name', '') or '',
                            last_name=row_data.get('last_name', '') or '',
                            company_name=row_data.get('company_name', '') or '',
                            phone=str(row_data.get('phone', '') or ''),
                            email=row_data.get('email', '') or '',
                            city=row_data.get('city', '') or '',
                            state=row_data.get('state', '') or '',
                            country=row_data.get('country', '') or '',
                            pin_code=str(row_data.get('pin_code', '') or ''),
                            gstin=row_data.get('gstin', '') or '',
                            pan=row_data.get('pan', '') or '',
                            customer_type=resolve_type(row_data.get('customer_type')),
                            address_line1=row_data.get('address_line1', '') or '',
                        )
                        created += 1
                    except Exception as e:
                        errors.append(f"Row {i}: {str(e)}")
            elif file.name.endswith('.csv'):
                decoded = file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(decoded))
                for i, row_data in enumerate(reader, start=2):
                    try:
                        Customer.objects.create(
                            first_name=row_data.get('first_name', ''),
                            last_name=row_data.get('last_name', ''),
                            company_name=row_data.get('company_name', ''),
                            phone=row_data.get('phone', ''),
                            email=row_data.get('email', ''),
                            city=row_data.get('city', ''),
                            state=row_data.get('state', ''),
                            country=row_data.get('country', ''),
                            pin_code=row_data.get('pin_code', ''),
                            gstin=row_data.get('gstin', ''),
                            pan=row_data.get('pan', ''),
                            customer_type=resolve_type(row_data.get('customer_type')),
                            address_line1=row_data.get('address_line1', ''),
                        )
                        created += 1
                    except Exception as e:
                        errors.append(f"Row {i}: {str(e)}")
            else:
                return Response({'error': 'Unsupported file format'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'message': f'{created} customers imported successfully',
            'created': created,
            'errors': errors,
        })

    @action(detail=False, methods=['get'], url_path='export/template')
    def export_template(self, request):
        fmt = request.query_params.get('format', 'xlsx')
        headers = ['first_name', 'last_name', 'company_name', 'address_line1', 'city', 'state', 'country', 'pin_code', 'phone', 'email', 'gstin', 'pan', 'customer_type']

        if fmt == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="customer_template.csv"'
            writer = csv.writer(response)
            writer.writerow(headers)
            return response

        wb = Workbook()
        ws = wb.active
        ws.title = 'Customers'
        ws.append(headers)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="customer_template.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def export(self, request):
        customers = self.filter_queryset(self.get_queryset())
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="customers.csv"'
        writer = csv.writer(response)
        writer.writerow(['first_name', 'last_name', 'company_name', 'city', 'state', 'phone', 'email', 'gstin', 'customer_type', 'is_active'])
        for c in customers:
            writer.writerow([c.first_name, c.last_name, c.company_name, c.city, c.state, c.phone, c.email, c.gstin, c.customer_type.name if c.customer_type else '', c.is_active])
        return response
